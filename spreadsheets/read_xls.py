# -*- coding: utf-8 -*-
''' Simple example that shows how to read from an Excel file.'''

__author__= "Luis C. Pérez Tato (LCPT)"
__copyright__= "Copyright 2025, LCPT"
__license__= "GPL"
__version__= "3.0"
__email__= "l.pereztato@ciccp.es"

from openpyxl import load_workbook
workbook = load_workbook(filename="xls_example.xlsx")

load_dict_sls= dict() # SLS loads.
load_dict_uls= dict() # ULS loads.
internal_force_codes= ['N', 'VY', 'VZ', 'MT', 'MY', 'MZ']
sheet = workbook[workbook.sheetnames[0]]
for row in sheet.iter_rows():
    lc= row[0].value
    if(lc):
        is_load_combination= isinstance(lc, int)
        if(is_load_combination):
            lc_data= dict()
            lc_data['code']= lc
            lc_title= row[1].value
            lc_data['title']= lc_title
            for i, int_force_code in enumerate(internal_force_codes):
                lc_data[int_force_code]= row[5+i].value*1e3
            is_uls= (lc_title[-3:]=='ELU')
            if(is_uls):
                load_dict_uls[lc]= lc_data
            else:
                load_dict_sls[lc]= lc_data

load_comb_dict= {'SLS': load_dict_sls, 'ULS': load_dict_uls}
# print(load_comb_dict)
